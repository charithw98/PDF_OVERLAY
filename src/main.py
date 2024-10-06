import msoffcrypto
import openpyxl
import os
import re
import getpass

# passwords perdata and saldata 

ERROR_SUCCESS = 0
ERROR_FILE_NOT_FOUND = 1

# create temp data file from source.
def createTempDataFile(sourceFile, password, tempFile):
    print("Fn: createTempDataFile")
    #check if the source file is there
    if( not os.path.exists(sourceFile)):
        print("Error: Source file not found")
        return ERROR_FILE_NOT_FOUND # error 
    with open(sourceFile, 'rb') as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password=password)  # Provide the password
        # Decrypt the file and save it as a temporary file
        with open(tempFile, 'wb') as decrypted_file:
            office_file.decrypt(decrypted_file)
    return ERROR_SUCCESS

# create temp data files from source. The file list has
# source, password and temp file names. Will return temp 
# file list. If any file does not need a password, use False
def createTempDataFiles(FileList):
    print("Fn: createTempDataFile")
    # variable to return the saved file list.
    fileMap = []
    for currentFile in FileList:
        # Open the protected file using msoffcrypto
        # NO error handling <todo>
        # files without a password, not implemented. <todo> 
        with open(currentFile["source"], 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=currentFile["password"])  # Provide the password
            # Decrypt the file and save it as a temporary file
            with open(currentFile["destination"], 'wb') as decrypted_file:
                office_file.decrypt(decrypted_file)
        # add the created file name to the fileMap 
        fileData = {"source": currentFile["source"], "destination": currentFile["tempFileName"]}
        fileMap.append(fileData)
    #return the list of converted files.
    return fileMap

# Remove temp files.
# the file list is a python dict, that has two elememnts.
# name: string delete: True/False
# return the error count
def removeTempFiles(fileList):
    print("Fn: removeTempFiles")
    errorCount = 0
    for files in fileList:
        if(files["delete"] & os.path.exists(files["name"])):
            os.remove(files["name"])
            print(f"Temporary files deleted.")
        else:
            print(f"ERROR: File not found.")
            print(files["name"])
            errorCount = errorCount + 1
    return errorCount

# Read template file, and load the data 
def loadTemplateData(templateFile,sheetName):
    # variable to return data
    textOverlayList = [] 
    #check if the file path is valid
    if( not os.path.exists(templateFile)):
        print("Tempate file not found")
        return textOverlayList # error - empty
    #open template file, and load the sheet
    templateBook = openpyxl.load_workbook(templateFile)
    textOverlayDataSheet = templateBook[sheetName]
    print("Debug:",textOverlayDataSheet.dimensions, textOverlayDataSheet.max_row, textOverlayDataSheet.max_column)
    #go through every text overlay item
    for overlays in textOverlayDataSheet.rows:
        # stop if we reach an empty cell
        if(None == overlays[0].value):
            break
        # the cordinates has to be always numbers, skip others
        if(not (isinstance(overlays[2].value,int) and isinstance(overlays[3].value,int))):
            print("Warning: skip Row ID ", overlays[0].value)
            continue
        dataString = str(overlays[4].value)
        #user initiated end of loop.
        if("None" == dataString):
            break
        if(dataString.startswith('!<') and dataString.endswith('>') and len(dataString) > 3):
            #process the file data. Get all the data points to a list
            fileData = re.findall(r'<(.*?)>',dataString)
            #check the item 2, File locked
            isFileLocked = False
            if fileData[1] == "LOCKED=1":
                isFileLocked = True
            #save the extended data
            textOverlayList.append({"name": overlays[1].value, 
                                    "text":{ "string": None, "x": overlays[2].value, "y": overlays[3].value},
                                    "file" : {"name": fileData[0], "isLocked": isFileLocked, "sheet": fileData[2], "primeryKey": fileData[3], "value": fileData[4]}})
        else:
            #Save notmal text data
            textOverlayList.append({"name": overlays[1].value, "text":{ "string": dataString, "x": overlays[2].value, "y": overlays[3].value}})
    # Data store is done. return
    print("Fn: loadTemplateData")
    return textOverlayList

# Get the list of locked source files from text overlay list
def getLockedFileList(textOverlayList):
    fileNameList = []
    #go through each overlay
    for testOverlay in textOverlayList:
        text = testOverlay["text"]
        textString = text["string"]
        # Check if this is a direct text
        if textString == None:
            #indirect text, taken from a different file
            fileData = testOverlay["file"]
            fileName = fileData["name"]
            isFileLocked = fileData["isLocked"]
            print ((fileName not in fileNameList),isFileLocked, ((fileName not in fileNameList) and isFileLocked))
            if ((fileName not in fileNameList) and isFileLocked):
                fileNameList.append(fileName)
    return fileNameList


def OpenAllDataFiles(overlayDataList, tempFileList):
    dataFileList =[]
    for overlayData in overlayDataList:
        print(overlayData["name"])
        overlayText = overlayData["text"]
        if(None == overlayText["string"]):
            #this is a file
            file = overlayData["file"]
            fileName = file["name"]
            #check if this is a loacked file
            if not file["isLocked"]:
                print("Open File",fileName)
                dataFileList.append({"sourceFileName": fileName, "fileObject": None})
            else:
                #open the temp file
                for tempFile in tempFileList:
                    if tempFile["source"] == fileName:
                        print ("OpenF Temp File",tempFile["tempFileName"])
                        dataFileList.append({"sourceFileName": fileName, "fileObject": None})
    return dataFileList

def getOverlayData(overlayLineInformation, tempFileList):
    OverlayData = []
    OverlayData.append({"string": "Test String", "x": 100, "y": 250 })
    return OverlayData


# main program

#load the template
textOverlayList =  loadTemplateData("TEMPLATE.xlsx","Overlay")

#get the locked files to be used as source 
lockedSourceFileList = getLockedFileList(textOverlayList)

#create temp files for locked files
tempFileList = []
for lockedSourceFile in lockedSourceFileList:
    tempFileIndex = 1
    SourceFileName = lockedSourceFile
    TempFileName = "~temp~" + str(tempFileIndex) + ".xlsx"
    # Prompt the user for a password and store it in a variable
    password = getpass.getpass("Enter password for "+ SourceFileName + ": ")
    if ERROR_SUCCESS == createTempDataFile(SourceFileName,password,TempFileName):
        tempFileList.append({"source": SourceFileName, "tempFileName": TempFileName})

print(tempFileList)

dataFilesList = OpenAllDataFiles(textOverlayList,tempFileList)

for dataFile in dataFilesList:
    print(dataFile["sourceFileName"])

for overlayData in textOverlayList:
    getOverlayData(overlayData,dataFilesList)

# Open temp files and read
for tempFile in tempFileList:
    print ("Open temp file", tempFile["tempFileName"])
    workbook = openpyxl.load_workbook(tempFile["tempFileName"])
    #open active sheet
    sheet = workbook.active
    # Read the value from cell B6
    cell_value = sheet['B6'].value
    print(f"The value in B6 is: {cell_value}")

deleteFileList = []
for tempFile in tempFileList:
    deleteFileList.append({"name": tempFile["tempFileName"], "delete": True})

errorCount = removeTempFiles(deleteFileList)
print("Error Count = ", errorCount)